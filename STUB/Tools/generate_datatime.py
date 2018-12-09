import os
import time
import calendar
import random
from openpyxl import load_workbook
# -*- coding: UTF-8 -*-

WORKPATH=os.path.dirname(os.path.realpath(__file__))
DOCDIR=os.path.realpath(WORKPATH + "\..\Scenario")
DOCPATH=os.path.join(DOCDIR, "case1.xlsx")

print(DOCPATH)

wb = load_workbook(DOCPATH)
ws=wb["40000"]

print(time.localtime(calendar.timegm(time.gmtime()[0:3] + (0, 0, 0))))
epoch_sec=calendar.timegm(time.gmtime()[0:3] + (0, 0, 0))
increamental=0
while (increamental<86400):
    cur_time = time.gmtime(epoch_sec+increamental)
    #print(time.strftime("%Y-%m-%d %H:%M:%S", cur_time) )
    increamental = increamental + 1
    #ws["A" + str(increamental)] = time.strftime("%Y-%m-%d %H:%M:%S", cur_time)
    ws["A" + str(increamental)] = time.strftime("%H:%M:%S", cur_time)
    ws["B" + str(increamental)] = random.randint(4000, 20000)

wb.save(DOCPATH)

#print(time.time())

