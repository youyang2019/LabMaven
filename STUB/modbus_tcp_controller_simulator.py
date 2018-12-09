#!/usr/bin/env python
# -*- coding: utf_8 -*-
'''
Author     ：sammyjeep
Date Time  ：2018/12/02
Description：modbus test stub.
'''
import sys
import logging
import threading
import modbus_tk
import modbus_tk.defines as cst
import modbus_tk.modbus as modbus
import modbus_tk.modbus_tcp as modbus_tcp
import random
import time
import os
from openpyxl import load_workbook

TEST_DATA={}

WORKPATH = os.path.dirname(os.path.realpath(__file__))
DOCDIR = os.path.realpath(WORKPATH + "\Scenario")
DOCPATH = os.path.join(DOCDIR, "case1.xlsx")
wb = load_workbook(DOCPATH)
#cell_range = ws['A1':'C2']
#print(cell_range)
for sheetname in wb.sheetnames:
    #if ws.title().decode(encoding='UTF-8').isnumeric():
    if not sheetname.isnumeric():
        continue
    TEST_DATA[sheetname] = {}
    for row in wb[sheetname].iter_rows(min_row=1, max_col=2):
        TEST_DATA[sheetname][row[0].value]=row[1].value
print(TEST_DATA)

LOGGER = modbus_tk.utils.create_logger(name="console", record_format="%(message)s")
if __name__ == "__main__":
    try:
        # server里的address需要写的树莓派的IP和需要开放的端口，注意开放相应的端口
        SERVER = modbus_tcp.TcpServer(address="127.0.0.1", port=12600)
        LOGGER.info("running...")
        LOGGER.info("enter 'quit' for closing the server")
        # 服务启动
        SERVER.start()
        # 建立第一个从机
        SLAVE1 = SERVER.add_slave(1)
        for reg_addr in TEST_DATA.keys():
            SLAVE1.add_block('A' + reg_addr, cst.HOLDING_REGISTERS, int(reg_addr), 1)
        #SLAVE1.add_block('A', cst.HOLDING_REGISTERS, 0, 4)#地址0，长度4
        #SLAVE1.add_block('B', cst.HOLDING_REGISTERS, 4, 14)
        #SLAVE1.set_values('A', 0, [1, 9, 30000 ,2]) #改变在地址0处的寄存器的值
        #SLAVE1.set_values('B', 4, [1, 2, 3, 4, 5, 6, 12, 1232])     #改变在地址4处的寄存器的值
        # SLAVE1.add_block('A1', cst.HOLDING_REGISTERS, 40001, 1)  # 地址0，长度4
        # SLAVE1.add_block('A2', cst.HOLDING_REGISTERS, 40002, 1)  # 地址0，长度4
        # SLAVE1.add_block('A3', cst.HOLDING_REGISTERS, 40003, 1)  # 地址0，长度4
        # SLAVE1.add_block('A4', cst.HOLDING_REGISTERS, 40004, 1)  # 地址0，长度4
        # SLAVE1.add_block('A5', cst.HOLDING_REGISTERS, 40005, 1)  # 地址0，长度4
        # SLAVE1.add_block('A6', cst.HOLDING_REGISTERS, 40006, 1)  # 地址0，长度4
        # SLAVE1.add_block('A7', cst.HOLDING_REGISTERS, 40007, 1)  # 地址0，长度4
        # SLAVE1.add_block('A8', cst.HOLDING_REGISTERS, 40008, 1)  # 地址0，长度4
        # SLAVE1.add_block('A9', cst.HOLDING_REGISTERS, 40009, 1)  # 地址0，长度4
        # SLAVE1.add_block('A10', cst.HOLDING_REGISTERS, 40010, 1)  # 地址0，长度4
        # SLAVE1.add_block('A11', cst.HOLDING_REGISTERS, 40011, 1)  # 地址0，长度4
        # SLAVE1.add_block('A12', cst.HOLDING_REGISTERS, 40012, 1)  # 地址0，长度4
        # SLAVE1.add_block('A13', cst.HOLDING_REGISTERS, 40013, 1)  # 地址0，长度4
        # SLAVE1.add_block('A14', cst.HOLDING_REGISTERS, 40014, 1)  # 地址0，长度4

        while True:
            timestamp=time.strftime("%H:%M:%S")
            print(timestamp)
            for reg_addr in TEST_DATA.keys():
                print(TEST_DATA[reg_addr][timestamp])
                SLAVE1.set_values('A' + reg_addr, int(reg_addr), TEST_DATA[reg_addr][timestamp])  # 改变在地址0处的寄存器的值
            #SLAVE1.set_values('A1', 40001, [random.randint(2000, 20000)])  # 改变在地址0处的寄存器的值
            # SLAVE1.set_values('A2', 40002, [random.randint(2000, 20000)])  # 改变在地址0处的寄存器的值
            # SLAVE1.set_values('A3', 40003, [random.randint(2000, 20000)])  # 改变在地址0处的寄存器的值
            # SLAVE1.set_values('A4', 40004, [random.randint(2000, 20000)])  # 改变在地址0处的寄存器的值
            # SLAVE1.set_values('A5', 40005, [random.randint(2000, 20000)])  # 改变在地址0处的寄存器的值
            # SLAVE1.set_values('A6', 40006, [random.randint(2000, 20000)])  # 改变在地址0处的寄存器的值
            # SLAVE1.set_values('A7', 40007, [random.randint(2000, 20000)])  # 改变在地址0处的寄存器的值
            # SLAVE1.set_values('A8', 40008, [random.randint(2000, 20000)])  # 改变在地址0处的寄存器的值
            # SLAVE1.set_values('A9', 40009, [random.randint(2000, 20000)])  # 改变在地址0处的寄存器的值
            # SLAVE1.set_values('A10', 40010, [random.randint(2000, 20000)])  # 改变在地址0处的寄存器的值
            # SLAVE1.set_values('A11', 40011, [random.randint(2000, 20000)])  # 改变在地址0处的寄存器的值
            # SLAVE1.set_values('A12', 40012, [random.randint(2000, 20000)])  # 改变在地址0处的寄存器的值
            # SLAVE1.set_values('A13', 40013, [random.randint(2000, 20000)])  # 改变在地址0处的寄存器的值
            # SLAVE1.set_values('A14', 40014, [random.randint(2000, 20000)])  # 改变在地址0处的寄存器的值

            time.sleep(0.5)
        # CMD = sys.stdin.readline()
        # if CMD.find('quit') == 0:
        #     sys.stdout.write('bye-bye\r\n')
        #     break
        # else:
        #     sys.stdout.write("unknown command %s\r\n" % (args[0]))
    finally:
        SERVER.stop()